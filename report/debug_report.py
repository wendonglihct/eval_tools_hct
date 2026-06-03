"""DEBUG 模式：读取 Excel 报告，提取关键数据生成飞书卡片。

卡片结构：
- 第一排：标题 "多版本对比结果：model=DEBUG"
- 第一段表格：ver_map 和 tasks_dict 信息
- 第二段表格：使用 table 组件展示数据
"""

import os
import sys
import json
from typing import Any, Dict, List

from openpyxl import load_workbook

# 允许独立运行
_PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

from utils.paths import REPORTS_DIR

# 默认报告文件路径
DEFAULT_XLSX_PATH = os.path.join(REPORTS_DIR, "DT_HCT_20260602210332.xlsx")


def extract_table_data(xlsx_path: str) -> Dict[str, Any]:
    """从 Excel 提取表格数据。

    提取列：A(数据分类), C(数量), E(评测内容), J及之后(各版本数据)
    只提取大特性行及其直接下属的指标行（跳过小特性及其下属行）

    Returns:
        {version_labels: [...], rows: [{name, count, content, versions}]}
    """
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    # 列索引（openpyxl 从 1 开始）
    COL_A = 1   # 数据分类
    COL_C = 3   # 数量
    COL_E = 5   # 评测内容
    COL_J = 10  # 版本数据起始列

    # 提取版本标签（表头 J 列及之后）
    version_labels = []
    for col_idx in range(COL_J, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        version_labels.append(str(val) if val else "")

    # 提取数据行
    rows = []
    current_feature_name = ""
    in_small_feature = False  # 标记是否在小特性区域

    for row_idx in range(2, ws.max_row + 1):
        name_a = str(ws.cell(row=row_idx, column=COL_A).value or "").strip()

        # 判断行类型
        if name_a.startswith("--"):
            # 小特性行：跳过并进入小特性模式
            in_small_feature = True
            continue

        if name_a:
            # 大特性行：退出小特性模式，记录特性名称
            in_small_feature = False
            current_feature_name = name_a

        # 在小特性区域内的指标行也跳过
        if in_small_feature:
            continue

        # 提取数据
        count = ws.cell(row=row_idx, column=COL_C).value or ""
        content = ws.cell(row=row_idx, column=COL_E).value or ""

        if not content:
            continue

        # 提取版本数据
        versions = []
        for col_idx in range(COL_J, ws.max_column + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            versions.append(str(val) if val else "-")

        rows.append({
            "name": current_feature_name,
            "count": str(count) if count else "-",
            "content": str(content),
            "versions": versions,
        })

    wb.close()
    return {"version_labels": version_labels, "rows": rows}


def build_debug_card(table_data: Dict[str, Any], sheet_url: str = "") -> Dict[str, Any]:
    """构建飞书交互卡片 JSON 结构（使用 table 组件）。

    Args:
        table_data: 表格数据
        sheet_url: 结果表格链接
    """
    version_labels = table_data["version_labels"]
    rows = table_data["rows"]

    # 构建列定义
    columns = [
        {"name": "col_0", "display_name": "数据分类", "width": "auto", "data_type": "text"},
        {"name": "col_1", "display_name": "数量", "width": "auto", "data_type": "text"},
        {"name": "col_2", "display_name": "评测内容", "width": "auto", "data_type": "text"},
    ]
    for i, label in enumerate(version_labels):
        columns.append({
            "name": f"col_{i+3}",
            "display_name": label,
            "width": "auto",
            "data_type": "text",
        })

    # 构建行数据（输出所有大特性数据）
    table_rows = []
    for row in rows:
        row_data = {
            "col_0": row["name"],
            "col_1": row["count"],
            "col_2": row["content"],
        }
        for i, ver in enumerate(row["versions"]):
            row_data[f"col_{i+3}"] = ver
        table_rows.append(row_data)

    # 结果表格链接内容
    link_content = f"[结果表格链接]({sheet_url})" if sheet_url else "结果表格链接：待生成"

    elements = [
        # 标题
        {
            "tag": "markdown",
            "content": "**多版本对比结果**",
            "text_align": "left",
            "text_size": "normal_v2",
        },
        # 结果表格链接
        {
            "tag": "markdown",
            "content": link_content,
            "text_align": "left",
            "text_size": "normal",
        },
        # 简表标题
        {
            "tag": "markdown",
            "content": "**简表：**",
            "text_align": "left",
            "text_size": "normal",
        },
        # 表格
        {
            "tag": "table",
            "page_size": 10,
            "row_height": "low",
            "freeze_first_column": False,
            "header_style": {
                "text_align": "left",
                "text_size": "normal",
                "background_style": "none",
                "text_color": "default",
                "bold": True,
            },
            "columns": columns,
            "rows": table_rows,
        },
    ]

    return {
        "schema": "2.0",
        "config": {"update_multi": True},
        "body": {
            "direction": "vertical",
            "horizontal_spacing": "8px",
            "vertical_spacing": "8px",
            "horizontal_align": "left",
            "vertical_align": "top",
            "elements": elements,
        },
    }


def process_debug_report(xlsx_path: str = DEFAULT_XLSX_PATH, sheet_url: str = "") -> Dict[str, Any]:
    """DEBUG 模式主处理：读取 xlsx 文件，生成飞书表格卡片。

    Args:
        xlsx_path: Excel 文件路径，默认 outputs/reports/DT_HCT_20260602210332.xlsx
        sheet_url: 结果表格链接

    Returns:
        飞书卡片 JSON 结构
    """
    table_data = extract_table_data(xlsx_path)
    return build_debug_card(table_data, sheet_url)


if __name__ == "__main__":
    card = process_debug_report()
    print(f"数据行数: {len(card['body']['elements'][2]['rows'])}")
    # 保存卡片 JSON
    json_path = os.path.join(REPORTS_DIR, "debug_card.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(card, f, ensure_ascii=False, indent=2)
    print(f"卡片 JSON: {json_path}")