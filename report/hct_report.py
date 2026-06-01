import os
import re
import sys
import json
import operator
import subprocess
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# 允许直接 `python report/hct_report.py` 运行（debugger 场景）
_PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

from report.config_loader import load_hct_feature_config, get_config_path, load_mode_routing
from utils.paths import TASKS_DIR, REPORTS_DIR, ensure_output_dirs as _ensure_output_dirs


class TaskLoader:
    """pncops 下载缓存 + PASS 过滤 + meta 拼装。

    - 下载缓存到 outputs/tasks/task_<id>.json；同 task_id 复用不重复下载
    - PASS / EVAL FAIL 之外的场景直接丢弃
    - 把 sim 顶层若干字段（tenant / task_hardware_type 等）注入 meta，让 DSL 可访问
      `scenario_object.tenant : "hct_astra"` 这类条件
    """

    # PASS / EVAL FAIL 都视为可用样本；其他 (Error/Timeout/...) 直接丢弃
    PASS_STATUS = {"Pass", "EVAL FAIL"}

    # sim 顶层这些字段会被合并到 meta，使 DSL `scenario_object.<field>` 可访问
    SIM_LEVEL_FIELDS = (
        "tenant", "task_id", "task_hardware_type", "task_billing_pool",
        "data_billing_type", "agent_name", "agent_queue_name",
        "scenario_start_time", "scenario_execution_duration", "task_create_time",
    )

    @staticmethod
    def load_raw(task_id):
        """读取（必要时下载）task_<id>.json 原始内容。"""
        _ensure_output_dirs()
        file_path = os.path.join(TASKS_DIR, f"task_{task_id}.json")
        if not os.path.exists(file_path):
            cmd = ["pncops", "task-detail", "--task_id", str(task_id), "-o", file_path]
            subprocess.run(cmd, check=True)
        with open(file_path, "r", encoding="utf-8") as f:
            return json.load(f)

    @classmethod
    def extract_records(cls, task_id):
        """转换为 {scenario_id, meta, evaluation_result} 列表（已 PASS 过滤 + meta 注入）。"""
        data = cls.load_raw(task_id)
        records = []
        for sim in data.get("es_sim", []):
            if sim.get("scenario_execution_result") not in cls.PASS_STATUS:
                continue
            raw = sim.get("scenario_object", {}).get("raw", {})
            meta = dict(raw.get("meta", {}) or {})
            for k in cls.SIM_LEVEL_FIELDS:
                if k in sim and k not in meta:
                    meta[k] = sim[k]
            records.append(
                {
                    "scenario_id": sim.get("scenario_id"),
                    "meta": meta,
                    "evaluation_result": sim.get("evaluation_result", {}) or {},
                }
            )
        return records

    @classmethod
    def load_all(cls, task_ids):
        """按 task_ids 列表拉取所有任务，返回 {task_id: records}。重复 id 自动去重。"""
        out = {}
        for tid in task_ids:
            out[str(tid)] = cls.extract_records(str(tid))
        return out


class PairwiseAligner:
    """版本两两交集对齐器。

    每个对比版本与 base 取场景 id 交集，分别筛 base 和对比版本的 records 作为对齐子集。
    避免「全量交集导致统计偏小」（只要任一版本不跑某场景，全量交集就少一条）。
    """

    @staticmethod
    def align(task_to_records, base_task):
        aligned_task_records = {}
        aligned_base_records = {}
        base_records = task_to_records.get(base_task, [])
        base_ids = {r.get("scenario_id") for r in base_records}
        aligned_task_records[base_task] = list(base_records)
        aligned_base_records[base_task] = list(base_records)
        for task_id, records in task_to_records.items():
            if task_id == base_task:
                continue
            task_ids = {r.get("scenario_id") for r in records}
            pair_common = base_ids & task_ids
            aligned_task_records[task_id] = [r for r in records if r.get("scenario_id") in pair_common]
            aligned_base_records[task_id] = [r for r in base_records if r.get("scenario_id") in pair_common]
        return aligned_task_records, aligned_base_records


# ver_map 长度对齐：不足时按 M、N、...、Z、AA、AB... 顺序补齐
def _normalize_ver_map(ver_map, target_len):
    ver_map = list(ver_map or [])
    if len(ver_map) >= target_len:
        return ver_map[:target_len]
    idx = 0
    while len(ver_map) < target_len:
        q = idx
        label = ""
        while True:
            label = chr(ord("M") + (q % 26)) + label
            q = q // 26 - 1
            if q < 0:
                break
        ver_map.append(label)
        idx += 1
    return ver_map


class MetricExpr:
    """指标表达式 + per-record 取值。

    支持 3 种 agg 语义：
    - "sum"               : 累加 metric_value
    - "ave"               : 平均 metric_value（缺测在 Aggregator 层跳过）
    - "metric_result:<b>" : 按 metric_result==<b> 计数（每条 0/1）

    支持 expr 多 term：'A+B-C' → [(+,A), (+,B), (-,C)]，逐 term 取值后按符号合并。
    """

    @staticmethod
    def parse_terms(expr):
        text = str(expr or "").strip()
        if not text:
            return []
        norm = text.replace(" ", "")
        parts = re.findall(r"[+\-]?[^+\-]+", norm)
        terms = []
        for idx, part in enumerate(parts):
            if not part:
                continue
            sign = 1
            token = part
            if token[0] == "+":
                token = token[1:]
            elif token[0] == "-":
                sign = -1
                token = token[1:]
            elif idx > 0:
                sign = 1
            if token:
                terms.append((sign, token))
        return terms

    @staticmethod
    def metric_value(rec, metric_key):
        metric = rec.get("evaluation_result", {}).get(metric_key, {})
        v = metric.get("metric_value", 0)
        try:
            return float(v)
        except Exception:
            return 0.0

    @staticmethod
    def parse_result_agg(agg_mode):
        """agg='metric_result:<bool>' → True/False；否则 None。"""
        text = str(agg_mode or "").strip().lower()
        prefix = "metric_result:"
        if not text.startswith(prefix):
            return None
        target = text[len(prefix):].strip()
        if target in ("true", "1"):
            return True
        if target in ("false", "0"):
            return False
        return None

    @staticmethod
    def metric_result_equals(rec, metric_key, target):
        metric = rec.get("evaluation_result", {}).get(metric_key, {})
        raw = metric.get("metric_result", None)
        if isinstance(raw, bool):
            val = raw
        elif isinstance(raw, str):
            s = raw.strip().lower()
            if s in ("true", "1"):
                val = True
            elif s in ("false", "0"):
                val = False
            else:
                return 0.0
        else:
            return 0.0
        return 1.0 if val == target else 0.0

    @classmethod
    def record_term_value(cls, rec, metric_key, agg_mode):
        target = cls.parse_result_agg(agg_mode)
        if target is not None:
            return cls.metric_result_equals(rec, metric_key, target)
        return cls.metric_value(rec, metric_key)

    @classmethod
    def calc_on_record(cls, rec, expr, agg_mode="sum"):
        terms = cls.parse_terms(expr)
        if not terms:
            return 0.0
        return sum(sign * cls.record_term_value(rec, k, agg_mode) for sign, k in terms)

    @classmethod
    def calc_value_or_none(cls, rec, expr, agg_mode):
        """ave 模式专用：任一 term 的 metric_value 缺失/不可解析 → None。"""
        terms = cls.parse_terms(expr)
        if not terms:
            return 0.0
        if cls.parse_result_agg(agg_mode) is not None:
            return cls.calc_on_record(rec, expr, agg_mode)
        value = 0.0
        for sign, metric_key in terms:
            m = rec.get("evaluation_result", {}).get(metric_key)
            if not isinstance(m, dict) or "metric_value" not in m:
                return None
            raw = m.get("metric_value")
            if raw is None:
                return None
            try:
                value += sign * float(raw)
            except Exception:
                return None
        return value

    @classmethod
    def expr_label(cls, expr, agg_mode, suffix):
        terms = cls.parse_terms(expr)
        if not terms:
            return f"{agg_mode}()_{suffix}"
        pieces = []
        for idx, (sign, metric_key) in enumerate(terms):
            token = f"{agg_mode}({metric_key})_{suffix}"
            if idx == 0:
                pieces.append(token if sign > 0 else f"-{token}")
            else:
                pieces.append(("+" if sign > 0 else "-") + token)
        return "".join(pieces)

    @classmethod
    def formula_text(cls, agg_mode, numerator_expr, denominator_expr):
        num_comp = cls.expr_label(numerator_expr, agg_mode, "comp")
        num_base = cls.expr_label(numerator_expr, agg_mode, "base")
        if denominator_expr:
            den_base = cls.expr_label(denominator_expr, agg_mode, "base")
            return f"(({num_comp})-({num_base}))/({den_base})"
        return f"({num_comp})-({num_base})"


class RecordFilter:
    """记录级 filter DSL 编译器。

    支持两种语法：
    - `<key>.metric_value <op> <num>`  : 数值比较，op ∈ {<,<=,>,>=,==,!=}
    - `<key>.metric_result:<bool>`     : 结果布尔判断
    返回 Callable[[rec], bool]；为空/None 时返回 None（调用方可短路跳过）。
    """

    _VALUE_RE = re.compile(
        r'^\s*([A-Za-z0-9_]+)\s*\.\s*metric_value\s*(<=|>=|==|!=|<|>)\s*(-?\d+(?:\.\d+)?)\s*$'
    )
    _RESULT_RE = re.compile(
        r'^\s*([A-Za-z0-9_]+)\s*\.\s*metric_result\s*:\s*(true|false|0|1)\s*$',
        flags=re.IGNORECASE,
    )
    _OPS = {
        "<": operator.lt, "<=": operator.le,
        ">": operator.gt, ">=": operator.ge,
        "==": operator.eq, "!=": operator.ne,
    }

    @classmethod
    def compile(cls, filter_str):
        if not filter_str:
            return None
        text = str(filter_str).strip()
        if not text:
            return None

        m = cls._VALUE_RE.match(text)
        if m:
            key, op_text, num_text = m.group(1), m.group(2), m.group(3)
            cmp = cls._OPS[op_text]
            threshold = float(num_text)

            def _f(rec):
                metric = rec.get("evaluation_result", {}).get(key)
                if not isinstance(metric, dict) or "metric_value" not in metric:
                    return False
                v = metric.get("metric_value")
                if v is None:
                    return False
                try:
                    return cmp(float(v), threshold)
                except Exception:
                    return False
            return _f

        m = cls._RESULT_RE.match(text)
        if m:
            key, tgt = m.group(1), m.group(2).lower()
            target = tgt in ("true", "1")

            def _f(rec):
                return MetricExpr.metric_result_equals(rec, key, target) > 0.5
            return _f

        raise ValueError(
            f"无法解析 filter: {filter_str!r}（支持 '<key>.metric_value <op> <num>' 或 '<key>.metric_result:<bool>'）"
        )


# 旧函数保留为薄壳
class Aggregator:
    """聚合器：把 records 集合按 (agg, expr) 折叠成 stats 字典。

    - selected_metric(records, expr, agg)
        单一 task 的 expr 聚合值。ave 模式跳缺测；sum/metric_result:* 累加。
    - per_task_stats(selected_records, agg, expr, filter=None)
        多 task 字典 → {task_id: {count, metric_sum, records}}。
        record_filter 仅过滤参与 metric_sum 的记录；count 保留全量。
    - per_tag_stats(records_by_task, extractor, agg, expr, task_ids, filter=None)
        按 extractor 把每个 record 分桶到一个或多个 tag；返回
        (value_stats[tag][task_id], other_stats[task_id])
        其中 count = 命中场景数（用于"数量"列），valid_count = 进入 metric_sum 的记录数
        （ave 归一化用）。filter 未通过的记录仅 count，不进 sum/valid_count。
    """

    @staticmethod
    def init_task_stats(task_ids):
        return {task_id: {"count": 0, "metric_sum": 0.0} for task_id in task_ids}

    @staticmethod
    def selected_metric(records, expr, agg_mode):
        if not records:
            return 0.0
        if str(agg_mode).strip().lower() == "ave":
            values = [v for v in (MetricExpr.calc_value_or_none(r, expr, agg_mode) for r in records) if v is not None]
            return sum(values) / len(values) if values else 0.0
        return sum(MetricExpr.calc_on_record(r, expr, agg_mode) for r in records)

    @classmethod
    def per_task_stats(cls, selected_records, agg_mode, metric_expr, record_filter=None):
        stats = {}
        for task_id, records in selected_records.items():
            filt_records = [r for r in records if record_filter(r)] if record_filter else records
            stats[task_id] = {
                "count": len(records),
                "metric_sum": cls.selected_metric(filt_records, metric_expr, agg_mode),
                "records": records,
            }
        return stats

    @classmethod
    def per_tag_stats(cls, records_by_task, extractor, agg_mode, metric_expr, task_ids, record_filter=None):
        value_stats = {}
        other_stats = cls.init_task_stats(task_ids)
        is_ave = str(agg_mode).strip().lower() == "ave"

        def _accum(target, metric_val, valid):
            target["count"] += 1
            if is_ave:
                if valid:
                    target["metric_sum"] += metric_val
                    target["valid_count"] = target.get("valid_count", 0) + 1
            else:
                target["metric_sum"] += metric_val

        for task_id, recs in records_by_task.items():
            for rec in recs:
                passes = record_filter(rec) if record_filter else True
                if not passes:
                    valid = False
                    metric_val = 0.0
                elif is_ave:
                    opt = MetricExpr.calc_value_or_none(rec, metric_expr, agg_mode)
                    valid = opt is not None
                    metric_val = opt if valid else 0.0
                else:
                    metric_val = MetricExpr.calc_on_record(rec, metric_expr, agg_mode)
                    valid = True
                raw_values = extractor(rec.get("meta", {}))
                values = sorted({v for v in raw_values if v})
                if not values:
                    _accum(other_stats[task_id], metric_val, valid)
                    continue
                for value in values:
                    if value not in value_stats:
                        value_stats[value] = cls.init_task_stats(task_ids)
                    _accum(value_stats[value][task_id], metric_val, valid)
        if is_ave:
            for stats in value_stats.values():
                for tid in stats:
                    vc = stats[tid].get("valid_count", 0)
                    stats[tid]["metric_sum"] = (stats[tid]["metric_sum"] / vc) if vc else 0.0
            for tid in other_stats:
                vc = other_stats[tid].get("valid_count", 0)
                other_stats[tid]["metric_sum"] = (other_stats[tid]["metric_sum"] / vc) if vc else 0.0
        return value_stats, other_stats


# 按 mode 从 modes.json 路由到具体的特性 JSON（未命中走 default_config）
def _feature_configs(mode=None):
    """按 mode 从 configs/modes.json 选择特性 JSON 并加载。
    - 命中 mode_to_config → 用映射文件
    - 未命中（含 None） → 用 default_config
    新 mode 在 modes.json 里追加即可，无需改代码。
    返回结构：
      [{name, sql, predicate, primary_result, results,
        small_features:[{field_sql, extractor, enabled}, ...]}, ...]
    """
    routing = load_mode_routing()
    default = routing.get("default_config", "hct_report_config.json")
    mapping = routing.get("mode_to_config", {}) or {}
    name = mapping.get(mode or "", default)
    return load_hct_feature_config(get_config_path(name))


class ValueFormatter:
    """单元格数值/差值/分母列的格式化集合。

    所有方法皆为纯函数，集中收编：
    - value(v, fmt)        : 单值 → 字符串，按 int/float/% 三种格式
    - raw(v, fmt)          : numerator 展示规则；float→3 位小数，其余→int
    - diff(delta, denom, use_denom, fmt) : 差值，有 denom 时换算成百分比
    - denominator_column(stats_by_task, task_order, use_denom):
                             基数(总数)列；多任务用 '/' 拼接
    """

    @staticmethod
    def value(value, value_format):
        fmt = str(value_format or "float").lower()
        if fmt == "%":
            return f"{value:.2f}%"
        if fmt == "int":
            return str(int(round(value)))
        return f"{value:.3f}"

    @staticmethod
    def raw(value, value_format):
        fmt = str(value_format or "float").lower()
        try:
            v = float(value)
        except Exception:
            v = 0.0
        if fmt == "float":
            return f"{v:.3f}"
        return str(int(round(v)))

    @staticmethod
    def diff(delta, denominator, use_denominator, value_format):
        if not use_denominator:
            return ValueFormatter.value(delta, value_format)
        if denominator == 0:
            if delta == 0:
                return ValueFormatter.value(0.0, value_format)
            return "N/A"
        ratio_percent = (delta / denominator) * 100
        return ValueFormatter.value(ratio_percent, value_format)

    @staticmethod
    def denominator_column(version_denominator_stats, task_order, use_denominator):
        """基数(总数)列：百分比指标按版本顺序列出 denominator，'/' 拼接。"""
        if not use_denominator:
            return "-"
        pieces = []
        for task_id in task_order:
            den = version_denominator_stats.get(task_id, {}).get("metric_sum", 0)
            try:
                pieces.append(f"{int(den)}")
            except Exception:
                pieces.append("0")
        return "/".join(pieces) if pieces else "-"


# 旧函数保留为薄壳，向后兼容（便于回滚 + 减少调用点扰动）
class RowComposer:
    """Excel 一行单元格的组装。

    输入：场景级 stats 字典 + 元信息（数据分类/sql/阈值/note 等）。
    输出：与表头一一对应的 list（含 9 个固定列 + N 个版本列）。

    版本列规则：
    - base 列：`{base_full}({/aligned/'/' 拼接})` （全量参考值 + 多对比版本对齐 base 值）
    - 对比列：`{value}({diff})` （版本聚合值 + 与 base 的差值，差值供着色器解析）
    """

    @staticmethod
    def build(
        data_class,
        sql_text,
        formula_text,
        result_name,
        note,
        rollback_threshold,
        optimize_threshold,
        task_order,
        stats,
        base_ref_stats,
        base_denominator_stats,
        version_denominator_stats,
        use_denominator,
        value_format,
    ):
        base_task = task_order[0]
        base_sum = stats.get(base_task, {}).get("metric_sum", 0.0)
        ref_task_for_count = task_order[1] if len(task_order) > 1 else base_task
        base_count = base_ref_stats.get(ref_task_for_count, {}).get(
            "count", stats.get(base_task, {}).get("count", 0)
        )
        denominator_cell = ValueFormatter.denominator_column(
            version_denominator_stats, task_order, use_denominator
        )
        row = [
            data_class, sql_text, base_count, formula_text, result_name,
            rollback_threshold, optimize_threshold, note, denominator_cell,
        ]
        for task_id in task_order:
            s = stats.get(task_id, {"count": 0, "metric_sum": 0.0})
            if task_id == base_task:
                base_full = ValueFormatter.raw(s.get("metric_sum", 0.0), value_format)
                compare_tasks = task_order[1:]
                aligned_pieces = [
                    ValueFormatter.raw(
                        base_ref_stats.get(t, {"metric_sum": base_sum}).get("metric_sum", 0.0),
                        value_format,
                    )
                    for t in compare_tasks
                ]
                row.append(f"{base_full}({'/'.join(aligned_pieces)})" if aligned_pieces else base_full)
                continue
            version_value = ValueFormatter.raw(s.get("metric_sum", 0.0), value_format)
            base_ref = base_ref_stats.get(task_id, {"metric_sum": base_sum})
            base_den = base_denominator_stats.get(task_id, {"metric_sum": 0.0})
            diff = ValueFormatter.diff(
                s["metric_sum"] - base_ref["metric_sum"],
                base_den["metric_sum"],
                use_denominator,
                value_format,
            )
            row.append(f"{version_value}({diff})")
        return row


# 旧函数薄壳
def _select_records_by_predicate(task_records, predicate):
    selected = {}
    for task_id, records in task_records.items():
        selected[task_id] = [r for r in records if predicate(r.get("meta", {}))]
    return selected


def _sum_counts(stats_by_task):
    return sum(v["count"] for v in stats_by_task.values())


class ExcelStyler:
    """Excel 表格样式：合并 / sheet 标题清洗 / 阈值着色。

    着色规则（rollback / optimize 都用 `<op><num>` 形式如 `<10`、`<-1%`）：
    - 单元格里若有 `(diff)` 括号，按 diff 数值着色；否则按整格
    - 双阈值：落 (optimize, rollback) 区间内不染；越靠近优秀色越绿；越靠近回退色越红
    - 单阈值：仅 rollback → 不满足染红；仅 optimize → 满足染绿
    """

    BETTER_COLORS = ["FFFFFFFF", "FFD9F5D6", "FF8EE085"]
    WORSE_COLORS = ["FFFFFFFF", "FFFAF1D1", "FFFAD355", "FFFBBFBC"]
    RED = "FFF54A45"
    GREEN = "FF34C724"
    WHITE = "FFFFFFFF"

    _THRESHOLD_RE = re.compile(r"(<=|>=|<|>)\s*(-?\d+\.?\d*)(?:%)?")
    _INNER_DIFF_RE = re.compile(r"\(([^)]+)\)")

    _OPS = {
        ">": operator.gt, "<": operator.lt,
        ">=": operator.ge, "<=": operator.le,
        "between": lambda v, lo, hi: lo < v < hi,
    }

    @staticmethod
    def merge_columns(ws, merge_ranges, col_indices=(1, 2, 3)):
        for start_row, end_row in merge_ranges:
            if end_row <= start_row:
                continue
            for col_idx in col_indices:
                ws.merge_cells(start_row=start_row, start_column=col_idx, end_row=end_row, end_column=col_idx)
                cell = ws.cell(row=start_row, column=col_idx)
                cell.alignment = Alignment(horizontal="left", vertical="center")

    @staticmethod
    def safe_sheet_title(raw_name, used_names):
        base = re.sub(r'[\[\]\:\*\?\/\\]', "_", str(raw_name or "DT_HCT"))
        base = base.strip() or "DT_HCT"
        if len(base) > 31:
            base = base[:31]
        title = base
        idx = 1
        while title in used_names:
            suffix = f"_{idx}"
            title = f"{base[:31 - len(suffix)]}{suffix}"
            idx += 1
        used_names.add(title)
        return title

    @classmethod
    def _parse_threshold(cls, cell_value):
        if not isinstance(cell_value, str):
            return None, None
        m = cls._THRESHOLD_RE.match(cell_value.strip())
        if not m:
            return None, None
        return m.group(1), float(m.group(2))

    @classmethod
    def _color_cell(cls, cell, color):
        cell.fill = PatternFill(fill_type="solid", start_color=color)

    @classmethod
    def apply_threshold_coloring(cls, ws, rollback_col=5, optimize_col=6, first_result_col=8):
        for row_idx in range(2, ws.max_row + 1):
            row_has_value = any(
                ws.cell(row=row_idx, column=c).value not in (None, "")
                for c in range(1, ws.max_column + 1)
            )
            if not row_has_value:
                continue

            rollback_op, rollback_number = cls._parse_threshold(ws.cell(row=row_idx, column=rollback_col).value)
            optimize_op, optimize_number = cls._parse_threshold(ws.cell(row=row_idx, column=optimize_col).value)

            for col_idx in range(first_result_col, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                raw = cell.value
                if raw is None or raw == "":
                    continue
                m = cls._INNER_DIFF_RE.search(str(raw))
                inner = m.group(1) if m else str(raw)
                try:
                    column_value = float(inner.replace("%", "").replace("=", "").strip())
                except Exception:
                    continue

                if rollback_number is not None and optimize_number is None:
                    if not cls._OPS[rollback_op](column_value, rollback_number):
                        cls._color_cell(cell, cls.RED)
                    continue
                if rollback_number is None and optimize_number is not None:
                    if cls._OPS[optimize_op](column_value, optimize_number):
                        cls._color_cell(cell, cls.GREEN)
                    continue
                if rollback_number is None and optimize_number is None:
                    continue

                low_value = min(rollback_number, optimize_number)
                high_value = max(rollback_number, optimize_number)

                if abs(column_value) < 0.001:
                    cls._color_cell(cell, cls.WHITE)
                elif cls._OPS["between"](column_value, low_value, high_value):
                    if (optimize_number > rollback_number) == cls._OPS["between"](column_value, 0.0, high_value):
                        palette = cls.BETTER_COLORS
                        ref = optimize_number
                    else:
                        palette = cls.WORSE_COLORS
                        ref = rollback_number
                    discrete = (ref - 0.0) / len(palette) if ref != 0 else 1
                    idx = int((column_value - 0.0) / discrete) if discrete else 0
                    idx = max(0, min(len(palette) - 1, idx))
                    cls._color_cell(cell, palette[idx])
                else:
                    if column_value == rollback_number:
                        cls._color_cell(cell, cls.RED)
                        continue
                    if column_value == optimize_number:
                        cls._color_cell(cell, cls.GREEN)
                        continue
                    if rollback_number <= optimize_number:
                        cls._color_cell(cell, cls.GREEN if cls._OPS[rollback_op](column_value, rollback_number) else cls.RED)
                    else:
                        cls._color_cell(cell, cls.GREEN if cls._OPS[optimize_op](column_value, optimize_number) else cls.RED)


def _stats_or_default(stats_map, key, task_ids):
    return stats_map.get(key, Aggregator.init_task_stats(task_ids))


class BundleBuilder:
    """把一个指标项(result_item) + 选择的记录集 → 用于渲染的统计 bundle。

    分离原因：
      - HCTReportBuilder 主流程只关心"我有这堆记录 + 这条指标定义，请给我可渲染数据"
      - Aggregator 只关心"按 task 聚合一个表达式"，不关心整体编排
      BundleBuilder 是这两层之间的编排层。
    """

    @classmethod
    def build(cls, selected_records, base_selected_records, small_features, result_item, task_ids):
        """small_features: 列表，每项 {field_sql, extractor, enabled}。
        返回 bundle 中 small_stats 为同长度列表，按下标对应每个小特性。"""
        agg_mode = result_item["agg"]
        numerator_expr = result_item["numerator"]
        denominator_expr = result_item.get("denominator")
        use_denominator = bool(denominator_expr)
        value_format = result_item.get("value_format", "%" if use_denominator else "float")
        record_filter = RecordFilter.compile(result_item.get("filter"))

        stats = Aggregator.per_task_stats(selected_records, agg_mode, numerator_expr, record_filter)
        base_ref_stats = Aggregator.per_task_stats(base_selected_records, agg_mode, numerator_expr, record_filter)
        if use_denominator:
            base_denominator_stats = Aggregator.per_task_stats(base_selected_records, agg_mode, denominator_expr, record_filter)
            version_denominator_stats = Aggregator.per_task_stats(selected_records, agg_mode, denominator_expr, record_filter)
        else:
            base_denominator_stats = Aggregator.init_task_stats(task_ids)
            version_denominator_stats = Aggregator.init_task_stats(task_ids)

        small_stats_list = [
            cls._build_small_stats(sf, selected_records, base_selected_records,
                                   agg_mode, numerator_expr, denominator_expr,
                                   use_denominator, record_filter, task_ids)
            for sf in small_features
        ]

        return {
            "result_item": result_item,
            "use_denominator": use_denominator,
            "formula_text": MetricExpr.formula_text(agg_mode, numerator_expr, denominator_expr),
            "value_format": value_format,
            "stats": stats,
            "base_ref_stats": base_ref_stats,
            "base_denominator_stats": base_denominator_stats,
            "version_denominator_stats": version_denominator_stats,
            "small_stats": small_stats_list,
        }

    @staticmethod
    def _build_small_stats(sf, selected_records, base_selected_records,
                           agg_mode, numerator_expr, denominator_expr,
                           use_denominator, record_filter, task_ids):
        ex = sf["extractor"]
        value_stats, other_stats = Aggregator.per_tag_stats(
            selected_records, ex, agg_mode, numerator_expr, task_ids, record_filter
        )
        base_value_stats, base_other_stats = Aggregator.per_tag_stats(
            base_selected_records, ex, agg_mode, numerator_expr, task_ids, record_filter
        )
        if use_denominator:
            base_value_den_stats, base_other_den_stats = Aggregator.per_tag_stats(
                base_selected_records, ex, agg_mode, denominator_expr, task_ids, record_filter
            )
            version_value_den_stats, version_other_den_stats = Aggregator.per_tag_stats(
                selected_records, ex, agg_mode, denominator_expr, task_ids, record_filter
            )
        else:
            base_value_den_stats = {}
            base_other_den_stats = Aggregator.init_task_stats(task_ids)
            version_value_den_stats = {}
            version_other_den_stats = Aggregator.init_task_stats(task_ids)
        return {
            "field_sql": sf.get("field_sql", ""),
            "enabled": sf.get("enabled", True),
            "value_stats": value_stats,
            "other_stats": other_stats,
            "base_value_stats": base_value_stats,
            "base_other_stats": base_other_stats,
            "base_value_den_stats": base_value_den_stats,
            "base_other_den_stats": base_other_den_stats,
            "version_value_den_stats": version_value_den_stats,
            "version_other_den_stats": version_other_den_stats,
        }


class HCTReportBuilder:
    """HCT 报告主流程编排。

    分层职责：
      - build_report：顶层入口，负责 Workbook 生命周期 + 跨 sheet 汇总 + 落盘
      - _render_sheet：单个 compare_key 的 sheet 渲染（特性循环 + 样式）
      - _render_feature：单个特性的 bundle 构建 + 大行 + 小特性遍历
      - _render_small_feature_block：小特性 value/other 行
      - _style_sheet：表头 / 大特性 / 列宽 / 阈值着色
    每层职责单一，外部直接调用 HCTReportBuilder.build_report。"""

    HEADER_FIXED = ["数据分类", "Sql语句", "数量", "分析逻辑", "评测内容",
                    "回退阈值", "优秀阈值", "注释", "基数(总数)"]
    COLUMN_WIDTHS = {"A": 26, "B": 30, "C": 6, "D": 30, "E": 16,
                     "F": 8, "G": 8, "H": 10, "I": 20, "J": 15}

    @classmethod
    def build_report(cls, tasks_dict, ver_map, file_name_suffix, mode="DT_HCT"):
        wb = Workbook()
        ws_default = wb.active
        used_sheet_names = set()
        has_written_sheet = False

        feature_configs = _feature_configs(mode)
        major_feature_names = {f["name"] for f in feature_configs}

        for compare_key, task_ids in tasks_dict.items():
            task_ids = [str(x) for x in task_ids]
            if not task_ids:
                continue
            if not has_written_sheet:
                ws = ws_default
                ws.title = ExcelStyler.safe_sheet_title(compare_key, used_sheet_names)
                has_written_sheet = True
            else:
                ws = wb.create_sheet(title=ExcelStyler.safe_sheet_title(compare_key, used_sheet_names))
            cls._render_sheet(ws, task_ids, ver_map, feature_configs, major_feature_names)

        if not has_written_sheet:
            ws_default.title = ExcelStyler.safe_sheet_title("DT_HCT", used_sheet_names)

        _ensure_output_dirs()
        xlsx_path = os.path.join(REPORTS_DIR, f"{mode}_{file_name_suffix}.xlsx")
        wb.save(xlsx_path)
        return xlsx_path

    # ---------- 单个 sheet ----------
    @classmethod
    def _render_sheet(cls, ws, task_ids, ver_map, feature_configs, major_feature_names):
        versions = _normalize_ver_map(ver_map, len(task_ids))
        header_task_ver = [f"{v}({t})" for v, t in zip(versions, task_ids)]
        all_rows = [cls.HEADER_FIXED + header_task_ver]
        merge_ranges = []
        major_feature_row_idx = set()
        next_row_idx = 2

        task_records_raw = TaskLoader.load_all(task_ids)
        task_records, base_ref_records = PairwiseAligner.align(task_records_raw, task_ids[0])

        for feature in feature_configs:
            next_row_idx = cls._render_feature(
                feature, task_ids, task_records, base_ref_records,
                all_rows, merge_ranges, major_feature_row_idx, next_row_idx,
            )

        for row in all_rows:
            ws.append(row)
        ExcelStyler.merge_columns(ws, merge_ranges, col_indices=(1, 2, 3))
        cls._style_sheet(ws, major_feature_row_idx, major_feature_names)

    # ---------- 单个特性 ----------
    @classmethod
    def _render_feature(cls, feature, task_ids, task_records, base_ref_records,
                        all_rows, merge_ranges, major_feature_row_idx, next_row_idx):
        selected_records = _select_records_by_predicate(task_records, feature["predicate"])
        base_selected_records = _select_records_by_predicate(base_ref_records, feature["predicate"])
        total_hits = sum(len(r) for r in selected_records.values()) + sum(
            len(r) for r in base_selected_records.values()
        )
        if total_hits == 0:
            return next_row_idx

        small_features = feature["small_features"]
        result_bundles = [
            BundleBuilder.build(
                selected_records=selected_records,
                base_selected_records=base_selected_records,
                small_features=small_features,
                result_item=result_item,
                task_ids=task_ids,
            )
            for result_item in feature["results"]
        ]
        primary_bundle = result_bundles[feature.get("primary_result", 0)]

        major_feature_row_idx.add(next_row_idx)
        feature_start_row = next_row_idx
        for idx, bundle in enumerate(result_bundles):
            result_item = bundle["result_item"]
            all_rows.append(
                RowComposer.build(
                    feature["name"] if idx == 0 else "",
                    feature["sql"] if idx == 0 else "",
                    bundle["formula_text"],
                    result_item["name"], result_item["note"],
                    result_item["rollback_threshold"], result_item["optimize_threshold"],
                    task_ids,
                    bundle["stats"], bundle["base_ref_stats"],
                    bundle["base_denominator_stats"], bundle["version_denominator_stats"],
                    bundle["use_denominator"], bundle["value_format"],
                )
            )
            next_row_idx += 1
        merge_ranges.append((feature_start_row, next_row_idx - 1))

        for sf_idx, sf_cfg in enumerate(small_features):
            if not sf_cfg.get("enabled", True):
                continue
            next_row_idx = cls._render_small_feature_block(
                feature, sf_idx, sf_cfg, result_bundles, primary_bundle, task_ids,
                all_rows, merge_ranges, next_row_idx,
            )
        return next_row_idx

    # ---------- 小特性 value + other ----------
    @classmethod
    def _render_small_feature_block(cls, feature, sf_idx, sf_cfg, result_bundles, primary_bundle,
                                    task_ids, all_rows, merge_ranges, next_row_idx):
        field_sql = sf_cfg.get("field_sql", "")
        primary_small = primary_bundle["small_stats"][sf_idx]
        primary_value_stats = primary_small["value_stats"]
        primary_other_stats = primary_small["other_stats"]

        value_order = sorted(
            primary_value_stats.keys(),
            key=lambda value: (-_sum_counts(primary_value_stats[value]), value),
        )

        for value in value_order:
            small_start_row = next_row_idx
            for idx, bundle in enumerate(result_bundles):
                result_item = bundle["result_item"]
                small = bundle["small_stats"][sf_idx]
                all_rows.append(
                    RowComposer.build(
                        f"--{value}" if idx == 0 else "",
                        f'{field_sql} = "{value}"' if idx == 0 else "",
                        bundle["formula_text"],
                        result_item["name"], result_item["note"],
                        result_item["rollback_threshold"], result_item["optimize_threshold"],
                        task_ids,
                        _stats_or_default(small["value_stats"], value, task_ids),
                        _stats_or_default(small["base_value_stats"], value, task_ids),
                        _stats_or_default(small["base_value_den_stats"], value, task_ids),
                        _stats_or_default(small["version_value_den_stats"], value, task_ids),
                        bundle["use_denominator"], bundle["value_format"],
                    )
                )
                next_row_idx += 1
            merge_ranges.append((small_start_row, next_row_idx - 1))

        if _sum_counts(primary_other_stats) > 0:
            other_start_row = next_row_idx
            for idx, bundle in enumerate(result_bundles):
                result_item = bundle["result_item"]
                small = bundle["small_stats"][sf_idx]
                all_rows.append(
                    RowComposer.build(
                        "--other" if idx == 0 else "",
                        f'{field_sql} = "other"' if idx == 0 else "",
                        bundle["formula_text"],
                        result_item["name"], result_item["note"],
                        result_item["rollback_threshold"], result_item["optimize_threshold"],
                        task_ids,
                        small["other_stats"], small["base_other_stats"],
                        small["base_other_den_stats"], small["version_other_den_stats"],
                        bundle["use_denominator"], bundle["value_format"],
                    )
                )
                next_row_idx += 1
            merge_ranges.append((other_start_row, next_row_idx - 1))
        return next_row_idx

    # ---------- 样式 ----------
    @classmethod
    def _style_sheet(cls, ws, major_feature_row_idx, major_feature_names):
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row_idx in major_feature_row_idx:
            data_class = ws.cell(row=row_idx, column=1).value
            if data_class in major_feature_names:
                ws.cell(row=row_idx, column=1).font = Font(color="FF0000", bold=True, size=14)
            ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="left", vertical="center")
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in (1, 2, 3):
                ws.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal="left", vertical="center")
        for col_letter, width in cls.COLUMN_WIDTHS.items():
            ws.column_dimensions[col_letter].width = width
        ExcelStyler.apply_threshold_coloring(ws, rollback_col=6, optimize_col=7, first_result_col=11)


if __name__ == "__main__":
    import time
    ver_map = ["base", "A", "B", "C"]
    tasks_dict = {"tmp": [1193157, 1193159, 1197924]}
    file_name_suffix = time.strftime("%Y%m%d%H%M%S", time.localtime())
    xlsx_path = HCTReportBuilder.build_report(tasks_dict, ver_map, file_name_suffix)
    print(f"xlsx: {xlsx_path}")

