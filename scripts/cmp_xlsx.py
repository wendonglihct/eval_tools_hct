"""对比两个 xlsx：值 / 合并区 / 列宽 / 字体加粗 / 单元格填充色。
用法: python scripts/cmp_xlsx.py baseline.xlsx new.xlsx
退出码 0 = 完全一致，非 0 = 有差异。
"""
import sys
from openpyxl import load_workbook


def dump_sheet(ws):
    rows = [
        tuple(c.value for c in row)
        for row in ws.iter_rows(values_only=False)
    ]
    merges = sorted(str(m) for m in ws.merged_cells.ranges)
    widths = {k: v.width for k, v in ws.column_dimensions.items() if v.width}
    fills = {}
    fonts = {}
    for row in ws.iter_rows():
        for c in row:
            if c.fill and c.fill.fgColor and c.fill.fgColor.rgb:
                rgb = c.fill.fgColor.rgb
                if rgb and rgb not in ("00000000", None):
                    fills[c.coordinate] = rgb
            if c.font and (c.font.bold or c.font.color):
                fonts[c.coordinate] = (c.font.bold, str(c.font.color.rgb) if c.font.color else None)
    return rows, merges, widths, fills, fonts


def cmp(a_path, b_path):
    wa = load_workbook(a_path)
    wb = load_workbook(b_path)
    sa, sb = wa.sheetnames, wb.sheetnames
    diffs = []
    if sa != sb:
        diffs.append(f"sheet 顺序/名字不同:\n  A={sa}\n  B={sb}")
    for name in sa:
        if name not in sb:
            continue
        ra, ma, wa_, fa, fta = dump_sheet(wa[name])
        rb, mb, wb_, fb, ftb = dump_sheet(wb[name])
        if ra != rb:
            # 找首个不同的 cell
            for i, (r1, r2) in enumerate(zip(ra, rb)):
                if r1 != r2:
                    for j, (v1, v2) in enumerate(zip(r1, r2)):
                        if v1 != v2:
                            diffs.append(f"[{name}] 单元格({i+1},{j+1}) A={v1!r} B={v2!r}")
                    break
            if len(ra) != len(rb):
                diffs.append(f"[{name}] 行数: A={len(ra)} B={len(rb)}")
        if ma != mb:
            diffs.append(f"[{name}] merge 不同:\n  仅A: {set(ma)-set(mb)}\n  仅B: {set(mb)-set(ma)}")
        if wa_ != wb_:
            for k in set(wa_) | set(wb_):
                if wa_.get(k) != wb_.get(k):
                    diffs.append(f"[{name}] 列宽 {k}: A={wa_.get(k)} B={wb_.get(k)}")
        if fa != fb:
            only_a = {k: v for k, v in fa.items() if fb.get(k) != v}
            only_b = {k: v for k, v in fb.items() if fa.get(k) != v}
            if only_a or only_b:
                # 仅列出最多 5 个
                for k in list(set(only_a) | set(only_b))[:5]:
                    diffs.append(f"[{name}] 填充色 {k}: A={fa.get(k)} B={fb.get(k)}")
        if fta != ftb:
            for k in list(set(fta) | set(ftb))[:5]:
                if fta.get(k) != ftb.get(k):
                    diffs.append(f"[{name}] 字体 {k}: A={fta.get(k)} B={ftb.get(k)}")
    if diffs:
        print("DIFF:")
        for d in diffs[:30]:
            print("  -", d)
        if len(diffs) > 30:
            print(f"  ... 共 {len(diffs)} 项差异")
        return 1
    print(f"OK: {a_path} == {b_path}")
    return 0


if __name__ == "__main__":
    sys.exit(cmp(sys.argv[1], sys.argv[2]))
